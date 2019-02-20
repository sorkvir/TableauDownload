from openpyxl import load_workbook
wb = load_workbook(filename = 'am.xlsx')
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
            for cell in row:
               print(getattr(cell, 'value', None))

for ws in wb.worksheets:
    print(ws)
'''
    #print(ws['B2'].value)
    for row in ws.values: # will throw merged cell ERROR, use getattr(item, 'value', None) to tell
       print(row) 
       for value in row:
         print(value)
'''


#xlrd too old

#otherwise use: win32com (not very fast but it works). This does allow you to get the formula. A tutorial can be found here and det
